import datetime
import random


import openpyxl
from kerykeion import CompositeAspects, KrInstance

from utils.db_api.db import BotDB


def get_value_find(workbook,sheet, rows, cols):
    # Задаем значение, которое нужно найти в таблице
    row_name = ''
    col_name = 'ali2'

    # Итерируемся по строкам и столбцам таблицы, пока не найдем нужное значение
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                               min_col=1, max_col=sheet.max_column):
        for cell in row:
            if row_name == '':
                if cell.value == rows:
                    # Найдено соответствующее значение - выводим координаты ячейки и значение
                    row_name = cell.coordinate
            else:
                if cell.value == cols:
                    # Найдено соответствующее значение - выводим координаты ячейки и значение
                    col_name = cell.coordinate
        else:
            continue

    cell_horizontal = sheet[col_name]
    cell_vertical = sheet[row_name]
    cell_value = sheet.cell(row=cell_horizontal.row, column=cell_vertical.column).value
    return cell_value


def get_radius_in_planet(p1_name,p2_name,aspect_degrees):
    if aspect_degrees in '0':
        path = f"data/0.xlsx"
    elif aspect_degrees in '60 120':
        path = f"data/60 120.xlsx"
    elif aspect_degrees in '90 180':
        path = f"data/90 180.xlsx"
    else:
        return 'None'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Лист1']
    result = get_value_find(workbook, sheet, p1_name, p2_name)
    if result == None or len(result) < 15:
        return 'None'
    return result


def get_moon_in_days(name_users,year_users,month_users,day_users,hours_user,minuts_users,lng_users,lat_users,
                         lng_last,lat_last,year_last,month_last,day_last,hour_last,minute_last):
    first = KrInstance(name_users, int(year_users), int(month_users), int(day_users), int(hours_user),
                       int(minuts_users), lng=float(lng_users), lat=float(lat_users))


    all_moon_array = []
    target_values = [0, 60, 120, 90, 180, 240, 270, 300]

    for i in range(0,24):
        for j in range(0,60):
            second = KrInstance(lng=float(lng_last), lat=float(lat_last), year=int(year_last), month=int(month_last),
                                day=int(day_last), hour=i, minute=j)

            name = CompositeAspects(first, second)
            numbers = []
            aspect_list = name.get_relevant_aspects()
            for planets_i in aspect_list:
                if 'Moon' in planets_i['p2_name']:
                    if 'House' in planets_i['p2_name'] or 'House' in planets_i['p1_name']:
                        continue
                    #print([planets_i['p2_name'], planets_i['p1_name'], planets_i['diff'], i, planets_i['aspect_degrees']])
                    for n in target_values:
                        if planets_i['diff'] < n + 0.1 and planets_i['diff'] > n - 0.1:
                            numbers.append([planets_i['p2_name'],planets_i['p1_name'],planets_i['diff'],i, planets_i['aspect_degrees']])

            try:
                closest_number = min(numbers, key=lambda x: min(abs(x[2] - target) for target in target_values))
            except:
                pass
            try:
                if all_moon_array[-1][1] == closest_number[1]:
                    continue
            except:
                pass
            try:
                all_moon_array.append(closest_number)
            except:
                pass
    return all_moon_array

def get_all_planet_in_days(name_users,year_users,month_users,day_users,hours_user,minuts_users,lng_users,lat_users,
                         lng_last,lat_last,year_last,month_last,day_last,hour_last,minute_last):
    first = KrInstance(name_users, int(year_users), int(month_users), int(day_users), int(hours_user),
                       int(minuts_users), lng=float(lng_users), lat=float(lat_users))
    target_values = [0, 60, 120, 90, 180, 240, 270, 300]
    second = KrInstance(lng=float(lng_last), lat=float(lat_last), year=int(year_last), month=int(month_last),
                        day=int(day_last))
    name = CompositeAspects(first, second)
    aspect_list = name.get_relevant_aspects()
    numbers = []
    for planets_i in aspect_list:
        if 'Moon' != planets_i['p2_name']and planets_i['p2_name'] in ['Sun','Mercury', 'Venus','Mars']:
            if 'House' in planets_i['p2_name'] or 'House' in planets_i['p1_name']:
                continue
            for n in target_values:
                #print(f"{planets_i['diff']} < {n + 2} | {planets_i['diff']} > {n - 2} ")
                if planets_i['diff'] < n + 0.1 and planets_i['diff'] > n - 0.1:
                    numbers.append(
                        [planets_i['p2_name'], planets_i['p1_name'], planets_i['diff'], planets_i['aspect_degrees']])

                    break



    #closest_number = min(numbers, key=lambda x: min(abs(x[2] - target) for target in target_values))
    #print(closest_number)
    return numbers

def get_planets_in_users(name_users,year_users,month_users,day_users,hours_user,minuts_users,lng_users,lat_users,
                         lng_last,lat_last,year_last,month_last,day_last,hour_last,minute_last):

    #get_radius_in_planet(planets_i['p1_name'], planets_i['p2_name'], str(planets_i['aspect_degrees']))
    all_moon_array = get_moon_in_days(name_users,year_users,month_users,day_users,hours_user,minuts_users,lng_users,lat_users,
                         lng_last,lat_last,year_last,month_last,day_last,hour_last,minute_last)
    all_planet_array = get_all_planet_in_days(name_users,year_users,month_users,day_users,hours_user,minuts_users,lng_users,lat_users,
                         lng_last,lat_last,year_last,month_last,day_last,hour_last,minute_last)
    print(all_planet_array)
    print(all_moon_array)
    return_string = ""
    for i in all_planet_array:
        try:
            result = get_radius_in_planet(i[0], i[1], str(i[3]))
            return_string += f"{result}\n\n"
        except:
            pass
    for i in all_moon_array:
        try:
            result = get_radius_in_planet(i[0], i[1], str(i[4]))

            if i[3]+3 < 5:
                continue
            if i[4] in [60, 90]:
                add_time = i[3]+3 + 2
            else:
                add_time = i[3]+3 + 3

            #print(f"<b>{i[3]}:00 - {add_time}:00</b>{result}\n\n")
            return_string += f"<b>{i[3]+3}:00 - {add_time}:00</b> {result}\n\n"
        except Exception as E:
            print(E)
    return return_string




def user_prog(answer_user,user_id):
    get_db_telegram = BotDB()
    user = get_db_telegram.get_users(user_id=user_id)
    print(user)
    users_info = ''

    date_of_birth = user['date_of_birth'].split('.')
    year_users, month_users, day_users = date_of_birth[2],date_of_birth[1],date_of_birth[0]

    time_of_birth = user['time_of_birth'].split(':')
    hours_user, minuts_users = time_of_birth[0], time_of_birth[1]
    try:
        place_of_birth = user['place_of_birth'].split(' ')
        lng_users, lat_users = place_of_birth[0], place_of_birth[1]
    except:
        lng_users, lat_users = '55.799584', '37.506832'


    city_now = user['city_now'].split(' ')
    lng_last, lat_last = city_now[0], city_now[1]
    if answer_user == 'today':
        time_last = datetime.datetime.now().strftime("%Y.%m.%d %H:%M").split(' ')
        days = time_last[0].split('.')
        year_last, month_last, day_last = days[0], days[1], days[2]
        ms = time_last[1].split(':')
        hour_last, minute_last = ms[0], ms[1]
        reslut = get_planets_in_users(name_users=user['name'], year_users=year_users, month_users=month_users,
                             day_users=day_users, hours_user=hours_user,
                             minuts_users=minuts_users, lng_users=lng_users, lat_users=lat_users,
                             lng_last=lng_last, lat_last=lat_last,
                             year_last=year_last, month_last=month_last, day_last=day_last,
                             hour_last="10", minute_last=minute_last)

        users_info = f'🧙 Прогноз на сегодня ({datetime.datetime.now().strftime("%Y.%m.%d")})\n\n {reslut}'

    elif answer_user == 'tomorrow':
        prog_str = datetime.datetime.now() + datetime.timedelta(days=1)
        time_last = prog_str.strftime("%Y.%m.%d %H:%M").split(' ')
        days = time_last[0].split('.')
        year_last, month_last, day_last = days[0], days[1], days[2]
        ms = time_last[1].split(':')
        hour_last, minute_last = ms[0], ms[1]
        reslut = get_planets_in_users(name_users=user['name'], year_users=year_users, month_users=month_users,
                                      day_users=day_users, hours_user=hours_user,
                                      minuts_users=minuts_users, lng_users=lng_users, lat_users=lat_users,
                                      lng_last=lng_last, lat_last=lat_last,
                                      year_last=year_last, month_last=month_last, day_last=day_last,
                                      hour_last="10", minute_last=minute_last)
        users_info = f'🧙 Прогноз на завтра ({prog_str.strftime("%Y-%m-%d") })\n\n {reslut}'

    elif answer_user == 'week':
        prog_today = datetime.datetime.now().strftime("%Y-%m-%d")
        prog_str = datetime.datetime.now() + datetime.timedelta(days=7)
        prog_str = datetime.datetime.now() + datetime.timedelta(days=7)
        time_last = prog_str.strftime("%Y.%m.%d %H:%M").split(' ')
        days = time_last[0].split('.')
        year_last, month_last, day_last = days[0], days[1], days[2]
        ms = time_last[1].split(':')
        hour_last, minute_last = ms[0], ms[1]
        reslut = get_planets_in_users(name_users=user['name'], year_users=year_users, month_users=month_users,
                                      day_users=day_users, hours_user=hours_user,
                                      minuts_users=minuts_users, lng_users=lng_users, lat_users=lat_users,
                                      lng_last=lng_last, lat_last=lat_last,
                                      year_last=year_last, month_last=month_last, day_last=day_last,
                                      hour_last="10", minute_last=minute_last)
        users_info = f'🧙 Прогноз на неделю ({prog_today} - {prog_str.strftime("%Y-%m-%d") })\n\n {reslut}'
    elif answer_user == 'month':
        prog_today = datetime.datetime.now().strftime("%Y-%m-%d")
        prog_str = datetime.datetime.now() + datetime.timedelta(days=31)
        prog_str = datetime.datetime.now() + datetime.timedelta(days=31)
        time_last = prog_str.strftime("%Y.%m.%d %H:%M").split(' ')
        days = time_last[0].split('.')
        year_last, month_last, day_last = days[0], days[1], days[2]
        ms = time_last[1].split(':')
        hour_last, minute_last = ms[0], ms[1]
        reslut = get_planets_in_users(name_users=user['name'], year_users=year_users, month_users=month_users,
                                      day_users=day_users, hours_user=hours_user,
                                      minuts_users=minuts_users, lng_users=lng_users, lat_users=lat_users,
                                      lng_last=lng_last, lat_last=lat_last,
                                      year_last=year_last, month_last=month_last, day_last=day_last,
                                      hour_last="10", minute_last=minute_last)
        users_info = f'🧙 Прогноз на месяц ({prog_today} - {prog_str.strftime("%Y-%m-%d") })\n\n {reslut}'
    elif answer_user == 'still':
        prog_today = datetime.datetime.now().strftime("%Y-%m-%d")
        prog_str = datetime.datetime.now()
        time_last = prog_str.strftime("%Y.%m.%d %H:%M").split(' ')
        days = time_last[0].split('.')
        year_last, month_last, day_last = days[0], days[1], days[2]
        ms = time_last[1].split(':')
        hour_last, minute_last = ms[0], ms[1]
        reslut = get_planets_in_users(name_users=user['name'], year_users=year_users, month_users=month_users,
                                      day_users=day_users, hours_user=hours_user,
                                      minuts_users=minuts_users, lng_users=lng_users, lat_users=lat_users,
                                      lng_last=lng_last, lat_last=lat_last,
                                      year_last=year_last, month_last=month_last, day_last=day_last,
                                      hour_last="10", minute_last=minute_last)
        users_info = f'🧙 Прогноз на сегодня ({datetime.datetime.now().strftime("%Y.%m.%d")})\n\n {reslut}' \
                     f'Следующий прогноз будет завтра в 10:00'
    else:
        try:
            prog_today = datetime.datetime.now().strftime("%Y-%m-%d")
            prog_str = datetime.datetime.now()
            time_last = prog_str.strftime("%Y.%m.%d %H:%M").split(' ')
            days = answer_user.split('.')
            year_last, month_last, day_last = days[2], days[1], days[0]
            ms = time_last[1].split(':')
            hour_last, minute_last = ms[0], ms[1]
            reslut = get_planets_in_users(name_users=user['name'], year_users=year_users, month_users=month_users,
                                          day_users=day_users, hours_user=hours_user,
                                          minuts_users=minuts_users, lng_users=lng_users, lat_users=lat_users,
                                          lng_last=lng_last, lat_last=lat_last,
                                          year_last=year_last, month_last=month_last, day_last=day_last,
                                          hour_last="10", minute_last=minute_last)
            users_info = f'🧙 Прогноз на ({answer_user})\n\n{reslut}'
        except:
            users_info = f'🧙 Прогноз не получился повторите попытку'

    return users_info


def get_img(date_now=None):
    from PIL import Image, ImageDraw, ImageFont

    im = Image.open('БОРДОВАЯ КАРТИНКА.jpg')
    if date_now == None:
        draw_text = ImageDraw.Draw(im)
        font = ImageFont.truetype("Comic Sans MS.ttf", 80)
        draw_text.text(
            (370, 450),
            datetime.datetime.now().strftime("%d-%m-%Y"),
            fill=('#ffffff'),
            font=font,
        )
    else:
        draw_text = ImageDraw.Draw(im)
        font = ImageFont.truetype("Comic Sans MS.ttf", 80)
        draw_text.text(
            (370, 450),
            date_now,
            fill=('#ffffff'),
            font=font,
        )
    im.save('БОРДОВАЯ КАРТИНКА2.jpg')
    return 'БОРДОВАЯ КАРТИНКА2.jpg'