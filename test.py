
import openpyxl
from kerykeion import CompositeAspects, KrInstance


def get_value_find(workbook,sheet, rows, cols):
    # Задаем значение, которое нужно найти в таблице
    row_name = ''
    col_name = 'ali2'

    # Итерируемся по строкам и столбцам таблицы, пока не найдем нужное значение
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                               min_col=1, max_col=sheet.max_column):
        for cell in row:
            if row_name == '':
                if cell.value == rows:
                    # Найдено соответствующее значение - выводим координаты ячейки и значение
                    row_name = cell.coordinate
            else:
                if cell.value == cols:
                    # Найдено соответствующее значение - выводим координаты ячейки и значение
                    col_name = cell.coordinate
        else:
            continue
            print(row_name)
    cell_horizontal = sheet[col_name]
    cell_vertical = sheet[row_name]
    cell_value = sheet.cell(row=cell_horizontal.row, column=cell_vertical.column).value
    return cell_value


def get_radius_in_planet(p1_name,p2_name,aspect_degrees):
    if aspect_degrees in '0':
        path = f"data/0.xlsx"
    elif aspect_degrees in '60 120':
        path = f"data/60 120.xlsx"
    elif aspect_degrees in '90 180':
        path = f"data/90 180.xlsx"
    else:
        return 'None'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Лист1']
    result = get_value_find(workbook, sheet, p1_name, p2_name)
    if result == None or len(result) < 15:
        return 'None'
    return result

def get_planets_in_users(name_users,year_users,month_users,day_users,hours_user,minuts_users,lng_users,lat_users,
                         lng_last,lat_last,year_last,month_last,day_last,hour_last,minute_last):
    first = KrInstance(name_users, int(year_users), int(month_users), int(day_users), int(hours_user), int(minuts_users), lng=float(lng_users), lat=float(lat_users))
    second = KrInstance(lng=float(lng_last), lat=float(lat_last), year=int(year_last), month=int(month_last), day=int(day_last), hour=int(hour_last), minute=int(minute_last))
    array_user = []
    name = CompositeAspects(first, second)
    aspect_list = name.get_relevant_aspects()
    for planets_i in aspect_list:

        try:
            result = get_radius_in_planet(planets_i['p1_name'], planets_i['p2_name'], str(planets_i['aspect_degrees']))

            print(planets_i['p1_name'],planets_i['p2_name'],planets_i['aspect_degrees'],int(planets_i['diff']))
        except:
            result = 'None'
        if result == 'None':
            continue
        #print(planets_i['p1_name'], planets_i['p2_name'], planets_i['aspect_degrees'])
        array_user.append(result)
    return array_user
get_planets_in_users(name_users='Maria',year_users="1979",month_users="11",day_users="15",hours_user="5",minuts_users="15",lng_users="48.023",lat_users="37.8022",
                     lng_last="48.023",lat_last="37.8022",year_last="2023",month_last="4",day_last="12",hour_last="12",minute_last="18")